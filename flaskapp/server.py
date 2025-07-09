# server.py
print("### SERVER.PY IMPORTED")

import ast
import os
import json
import time
import logging
import atexit
import tempfile
import shutil
import subprocess
import requests
import threading
from threading import Thread, Lock, Event
from google.oauth2 import service_account
from google.auth.transport.requests import Request
from datetime import datetime, date
from flask import Flask, request, jsonify, render_template, Response, stream_with_context
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from collections import deque
from logging.handlers import TimedRotatingFileHandler
from config import (
    LOG_FILE, DATA_PCS_FILE, DATA_MANUAL_FILE, PRIORITY_FILE,
    OPTIONS_FILE, FIXTURE_FILE, FIREBASE_CREDENTIAL_PATH,
    FIREBASE_PROJECT_ID, FIREBASE_COLLECTION, FIREBASE_ENABLED,
    SERIAL_PORT, SERIAL_BAUDRATE, DEBUG_MODE,
    VIEWER_PINGS_LOG_TEMPLATE, VIEWER_SESSIONS_FILE_TEMPLATE,
    OVERS_RULES, HOME_GROUND_IDENTIFIER, ADMIN_PASSWORD
)


os.environ["GRPC_DNS_RESOLVER"] = "native"

# Setup logger

os.makedirs(os.path.dirname(LOG_FILE), exist_ok=True)

logger = logging.getLogger('scoreboard')

# Ensure handler is only added once
if not any(isinstance(h, TimedRotatingFileHandler) and h.baseFilename == os.path.abspath(LOG_FILE) for h in logger.handlers):
    logger.setLevel(logging.DEBUG)

    file_handler = TimedRotatingFileHandler(
        LOG_FILE,
        when='midnight',
        interval=1,
        encoding='utf-8'
        # No backupCount — so no deletion
    )
    file_handler.suffix = "%Y-%m-%d"
    file_handler.setLevel(logging.DEBUG)

    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(formatter)

    logger.addHandler(file_handler)

logger.debug("########################### SCOREBOARD SERVER STARTED ###########################")



app = Flask(__name__)

# set up serial_manager (must be after logging setup)
from serial_manager import SerialManager, serial_mgr
from config import SERIAL_PORT, SERIAL_BAUDRATE


# set up Firebase publishing
from firebase_manager import FirebasePublisher, firebase_mgr, push_today_to_viewer_dates, init_firebase
from config import FIREBASE_ENABLED, FIREBASE_CREDENTIAL_PATH, FIREBASE_COLLECTION, FIREBASE_PROJECT_ID




# Default data structure: all fields blank except batting_team_score and overs
default_data = {
    "batting_team_score": "",
    "overs": "",  # Only used in manual mode
    "overs_bowled": "",
    "overs_remaining": "",
    "runs_required": "",
    "required_run_rate": "",
    "batting_team_name": "",
    "bowling_team_name": "",
    "bowling_team_score": "",
    "current_bowler_name": "",
    "current_bowler_score": "",
    "previous_bowler_name": "",
    "previous_bowler_score": "",
    "batter_1_name": "BATTER 1",
    "batter_1_score": "",
    "batter_1_balls": "",
    "batter_1_strike": "",
    "batter_2_name": "BATTER 2",
    "batter_2_score": "",
    "batter_2_balls": "",
    "batter_2_strike": "",
    "current_over_ball": "",
    "last_wicket_score": "",
    "last_wicket_batter": "",
    "last_wicket_batter_score": "",
    "last_wicket_dismissal": "",
    "last_wicket_bowler": "",
    "last_wicket_fielder": "",
    "dlt_score": "",
    "dlp_score": "",
    "target": "",  # used in manual mode
    "result": "",  # used post-match
    "last_updated_at": "",
    "ball_timeline": []
}
# Note: "startup_reset" is set manually for PCS only at runtime


# for viewer stats
client_sessions = {}

# Buffer for incoming PCS updates
pending_pcs_updates = []
last_update_time = None
pending_lock = Lock()
QUIET_TIME_SECONDS = 2.0  # quiet time to wait before processing pcs data
buffer_flush_timer = None


def init_data_file(file_path, default_data):
    """Ensure the file exists; write defaults if not."""
    if not os.path.exists(file_path):
        new_data = default_data.copy()
        new_data["last_updated_at"] = datetime.utcnow().isoformat() + "Z"
        with open(file_path, "w") as f:
            json.dump(new_data, f)


def init_priority_file():
    need_init = True
    if os.path.exists(PRIORITY_FILE):
        try:
            with open(PRIORITY_FILE, "r") as f:
                data = json.load(f)
            last_updated = data.get("updated_at", "")
            if last_updated:
                last_date = datetime.fromisoformat(last_updated.rstrip("Z")).date()
                if last_date == datetime.utcnow().date():
                    need_init = False
        except Exception:
            need_init = True
    if need_init:
        priority = {
            "active_source": "PCS",
            "updated_at": datetime.utcnow().isoformat() + "Z"
        }
        with open(PRIORITY_FILE, "w") as f:
            json.dump(priority, f)

def is_data_fresh(path):
    try:
        data = read_json(path)
        ts = data.get("last_updated_at", "")
        logger.debug(f"[Freshness] Checking timestamp: {ts}")
        if not ts:
            return False
        dt = datetime.fromisoformat(ts.rstrip("Z"))
        return dt.date() == datetime.utcnow().date()
    except Exception as e:
        logger.warning(f"[Freshness] Failed to parse timestamp: {e}")
        return False

def get_overs_for_fixture(fixture_type, division_or_cup):
    fixture_type = str(fixture_type or "").lower()
    division_or_cup = str(division_or_cup or "").lower()
    for rule in OVERS_RULES:
        match_type = rule.get("match_type", "").lower()
        division_sub = rule.get("division_contains", "").lower()
        logger.debug(f"[Fixture] Checking rule: {rule}")
        logger.debug(f"[Fixture] fixture_type='{fixture_type}' vs rule_type='{match_type}'")
        logger.debug(f"[Fixture] division_or_cup='{division_or_cup}' vs contains='{division_sub}'")

        if fixture_type == match_type:
            if not division_sub or division_sub in division_or_cup:
                logger.debug(f"[Fixture] Matched rule: {rule}")
                return rule["overs"]
    logger.debug("[Fixture] No matching rule found")
    return None


def auto_select_today_fixture():
    today = datetime.today().strftime("%d/%m/%Y")
    fixture_found = False

    options = {
        "home_team": "",
        "away_team": "",
        "show_batsmen": False,
        "overs_per_innings": 0,
        "last_updated_at": ""
    }

    logger.info(f"[Fixture] Checking for fixture on {today}")

    if os.path.exists(OPTIONS_FILE):
        try:
            with open(OPTIONS_FILE, 'r') as optf:
                existing = json.load(optf)
            options.update(existing)
            logger.info("[Fixture] Loaded existing options.json")
        except Exception as e:
            logger.warning(f"[Fixture] Failed to load options.json: {e}")

    if os.path.exists(FIXTURE_FILE):
        try:
            with open(FIXTURE_FILE, 'r') as f:
                fixtures = json.load(f)
            for fxt in fixtures:
                if fxt.get("date") == today:
                    fixture_found = True
                    options["home_team"] = fxt.get("home_team", "---")
                    options["away_team"] = fxt.get("away_team", "---")
                    logger.info(f"[Fixture] Found fixture: {options['home_team']} vs {options['away_team']}")

                    fixture_type = fxt.get("type", "")
                    division_or_cup = fxt.get("division_or_cup", "")

                    last_updated = options.get("last_updated_at", "")
                    updated_today = False
                    if last_updated:
                        try:
                            last_dt = datetime.fromisoformat(last_updated.rstrip("Z")).date()
                            updated_today = last_dt == datetime.utcnow().date()
                        except Exception as e:
                            logger.warning(f"[Fixture] Failed to parse last_updated_at: {e}")
                    logger.info(f"[Fixture] type: '{fixture_type}', division: '{division_or_cup}'")
                    overs = get_overs_for_fixture(fixture_type, division_or_cup)
                    if overs is not None:
                        if not updated_today:
                            options["overs_per_innings"] = overs
                            logger.info(f"[Fixture] Setting overs_per_innings to {overs} based on fixture rules")
                        else:
                            logger.info("[Fixture] Overs already updated today - skipping auto-set")
                    else:
                        logger.info("[Fixture] Fixture found but does not match any overs rule")

                    break
            else:
                logger.info("[Fixture] No fixture found for today in fixtures.json")
        except Exception as e:
            logger.error(f"[Fixture] Fixture check failed: {e}")
    else:
        logger.warning("[Fixture] fixtures.json not found")

    if not fixture_found:
        options["home_team"] = "---"
        options["away_team"] = "---"
        logger.info("[Fixture] No fixture found - resetting team names to '---'")

    options["last_updated_at"] = datetime.utcnow().isoformat() + "Z"

    try:
        with open(OPTIONS_FILE, 'w') as outf:
            json.dump(options, outf)
        logger.info("[Fixture] Updated options.json with fixture info")
    except Exception as e:
        logger.error(f"[Fixture] Failed to write options.json: {e}")



# Initialize files on startup 
init_data_file(DATA_PCS_FILE, default_data)
init_data_file(DATA_MANUAL_FILE, default_data)
init_priority_file()



# Helper functions to read/write data
def read_json(file_path, retries=3, delay=0.1):
    """Read a JSON file with optional retries to avoid partial read issues."""
    for attempt in range(retries):
        try:
            with open(file_path, "r") as f:
                data = json.load(f)
#                logger.debug(f"[read_json] Successfully read {file_path} on attempt {attempt + 1}")
                return data
        except Exception as e:
            logger.debug(f"[read_json] Attempt {attempt + 1} failed for {file_path}: {e}")
            if attempt < retries - 1:
                time.sleep(delay)
            else:
                logger.error(f"[read_json] Failed to read {file_path} after {retries} attempts: {e}")
    return {}


def write_json(file_path, data):
    """Write JSON data atomically to avoid partial writes."""
    try:
        dir_name = os.path.dirname(file_path)
        with tempfile.NamedTemporaryFile("w", delete=False, dir=dir_name) as tmp_file:
            json.dump(data, tmp_file)
            temp_name = tmp_file.name

        shutil.move(temp_name, file_path)
        logger.debug(f"[write_json] Successfully wrote {file_path}")
    except Exception as e:
        logger.error(f"[write_json] Failed to write {file_path}: {e}")


# helper for processing PCS data

def process_pending_updates():
    """Apply all pending PCS updates after quiet time."""
    global pending_pcs_updates, last_update_time, buffer_flush_timer

    with pending_lock:
        if not pending_pcs_updates:
            return

        combined_data = {}
        for update in pending_pcs_updates:
            combined_data.update(update)

        pending_pcs_updates.clear()
        last_update_time = None

    if not combined_data:
        return

    logger.info(f"[/update] Processing buffered PCS data (after quiet period): {combined_data}")

    data = read_json(DATA_PCS_FILE)

    previous_ovb_str = data.get("overs_bowled", "")
    previous_ovr_str = data.get("overs_remaining", "")

    # Apply all the new fields
    for key, value in combined_data.items():
        data[key] = value

    # Reset ball_timeline if OVB indicates new innings
    if combined_data.get("overs_bowled", "") == "0":
        if data.get("ball_timeline") != []:
            data["ball_timeline"] = []
            logger.info("[PCS Inning Reset] Overs bowled = 0 → Resetting ball_timeline")



    # Fallback logic: infer overs_bowled if missing and OVR changed
    if "overs_remaining" in combined_data and "overs_bowled" not in combined_data:
        try:
            new_ovr = int(combined_data["overs_remaining"])
            old_ovr = int(previous_ovr_str) if previous_ovr_str.isdigit() else None
            old_ovb = float(previous_ovb_str) if previous_ovb_str.replace(".", "", 1).isdigit() else None

            if old_ovr is not None and old_ovb is not None:
                if new_ovr == old_ovr - 1:
                    inferred_ovb = int(old_ovb) + 1
                    data["overs_bowled"] = str(inferred_ovb)
                    logger.info(f"[PCS Overs Fallback] Inferred OVB = {inferred_ovb} from OVR change: {old_ovr} -> {new_ovr}")
        except Exception as e:
            logger.warning(f"[PCS Overs Fallback] Failed to infer overs_bowled: {e}")

    if "startup_reset" in data:
        del data["startup_reset"]
        logger.info("[/update] Received first PCS data – startup_reset flag cleared")

    data["last_updated_at"] = datetime.utcnow().isoformat() + "Z"
    write_json(DATA_PCS_FILE, data)

    logger.info(f"[/update] PCS Data updated after buffering: {combined_data}")
    push_score_to_arduino()

    # Clear the buffer flush timer now that we've processed
    buffer_flush_timer = None


# -------------------------------
# Ball timeline update helper
# -------------------------------

MAX_BALLS = 48  # Number of balls + over markers to keep in ball_timeline

def update_ball_timeline_from_cov(new_cov):
    """Update the ball timeline inside data_pcs.json based on incoming COV field."""

    if new_cov is None:
        return

    ball_data = new_cov.strip()

    try:
        data = read_json(DATA_PCS_FILE)
    except Exception as e:
        logger.warning(f"[Update Ball Timeline] - Failed to read PCS file in ball timeline update: {e}")
        data = {}

    timeline = data.get('ball_timeline', [])
    if not isinstance(timeline, list):
        timeline = []

    if not ball_data:
        # COV " " received -> over has ended
        # Insert over marker if not already there
        if not timeline or timeline[-1] != '|':
            timeline.append('|')
    else:
        # COV with ball data -> replace current over
        balls = ball_data.split()

        # Find last over marker
        if '|' in timeline:
            last_over_index = len(timeline) - 1 - timeline[::-1].index('|')
            timeline = timeline[:last_over_index + 1]  # Keep up to last |
        else:
            timeline = []  # No over yet

        timeline.extend(balls)

    # Trim timeline if too long
    if len(timeline) > MAX_BALLS:
        timeline = timeline[-MAX_BALLS:]

    data['ball_timeline'] = timeline

    logger.info(f"[Update Ball Timeline] - Updated ball timeline: {' '.join(timeline)}")

    try:
        write_json(DATA_PCS_FILE, data)
    except Exception as e:
        logger.warning(f"[Update Ball Timeline] - Failed to write PCS file in ball timeline update: {e}")



def push_score_to_arduino():
    if serial_mgr is None:
        logger.error("[Serial] Cannot send - SerialManager not initialized.")
        return

    priority = read_json(PRIORITY_FILE)
    source = priority.get("active_source", "PCS")
    data_file = DATA_PCS_FILE if source == "PCS" else DATA_MANUAL_FILE
    data = read_json(data_file)


    score_str = data.get("batting_team_score", "0/0")
    # total, wkts = score_str.split('/') if '/' in score_str else ("0", "0")
    parts = score_str.split('/')
    if len(parts) == 2:
        total, wkts = parts
    elif len(parts) == 1:
        total, wkts = parts[0], ""  # send blank wickets if not provided
    else:
        total, wkts = "0", ""

    overs = data.get("overs_bowled", "0") if source == "PCS" else data.get("overs", "0")
    batsa = data.get("batter_1_score", "0")
    batsb = data.get("batter_2_score", "0")
    # Determine target based on mode
    if source == "PCS":
        bowling_score = data.get("bowling_team_score", "0")
        target = str(int(bowling_score) + 1) if bowling_score.isdigit() else "0"
    else:
        target = data.get("target", "0")

    # Prepare minimal dictionary for serial output
    score_data = {
        "total": int(total) if total.isdigit() else 0,
        "wickets": int(wkts) if wkts.isdigit() else None,
        "overs": int(float(overs)) if overs.replace('.', '', 1).isdigit() else 0,
        "batsa": int(batsa) if str(batsa).isdigit() else 0,
        "batsb": int(batsb) if str(batsb).isdigit() else 0,
        "target": int(target) if str(target).isdigit() else 0
    }

    try:
        serial_mgr.send_score(score_data)
        logger.info(f"Push score to Arduino: {score_data}, mode={source}")
    except Exception as e:
        logger.exception(f"[Push to Arduino] Failed to send to Arduino: {e}")
        

    # Push the score and mode to Firebase
    if firebase_enabled and firebase_mgr:
        firebase_mgr.publish("data", data)
        firebase_mgr.publish("priority", priority)
        # Push options to Firebase (team names, overs, etc.)
        try:
            options = read_json(OPTIONS_FILE)
            firebase_mgr.publish("options", options)
        except Exception as e:
            logger.warning(f"[push_to_arduino] Failed to read or publish options: {e}")





# Endpoints to serve data


@app.route('/data_pcs', methods=['GET'])
def data_pcs():
    return jsonify(read_json(DATA_PCS_FILE))

@app.route('/data_manual', methods=['GET'])
def data_manual():
    return jsonify(read_json(DATA_MANUAL_FILE))



@app.route('/data', methods=['GET'])
def data():
    priority = read_json(PRIORITY_FILE)
    active_source = priority.get("active_source", "PCS")
    data = read_json(DATA_PCS_FILE if active_source == "PCS" else DATA_MANUAL_FILE)

    return jsonify(data)






# new version for raw data from esp32
@app.route('/update', methods=['POST'])
def update_pcs():
    global pending_pcs_updates, last_update_time, buffer_flush_timer

    incoming = request.json
    logger.info(f"[/update] Incoming raw data: {incoming}")

    new_data = {}

    code_map = {
        "COV": "current_over_ball",
        "OVB": "overs_bowled",
        "OVR": "overs_remaining",
        "RRQ": "runs_required",
        "RRR": "required_run_rate",
        "BTN": "batting_team_name",
        "BTS": "batting_team_score",
        "FTN": "bowling_team_name",
        "FTS": "bowling_team_score",
        "B1N": "batter_1_name",
        "B1S": "batter_1_score",
        "B1B": "batter_1_balls",
        "B1K": "batter_1_strike",
        "B2N": "batter_2_name",
        "B2S": "batter_2_score",
        "B2B": "batter_2_balls",
        "B2K": "batter_2_strike",
        "F1N": "current_bowler_name",
        "F1S": "current_bowler_score",
        "F2N": "previous_bowler_name",
        "F2S": "previous_bowler_score",
        "LWK": "last_wicket_score",
        "LWN": "last_wicket_batter",
        "LWS": "last_wicket_batter_score",
        "LWD": "last_wicket_dismissal",
        "LWB": "last_wicket_bowler",
        "LWF": "last_wicket_fielder",
        "DLT": "dlt_score",
        "DLP": "dlp_score"
    }

    known_prefixes = list(code_map.keys())

    now = time.time()
    last_update_time = now  # Always update timestamp when something arrives

    if "data" in incoming:
        ble_string = incoming["data"]
        cursor = 0

        while cursor < len(ble_string):
            matched = False
            for prefix in known_prefixes:
                if ble_string.startswith(prefix, cursor):
                    cursor += len(prefix)
                    start = cursor
                    while cursor < len(ble_string) and not any(ble_string.startswith(p, cursor) for p in known_prefixes):
                        cursor += 1
                    value = ble_string[start:cursor]
                    field_name = code_map[prefix]
                    new_data[field_name] = value

                    logger.debug(f"[/update] Matched {prefix} -> {field_name}: {value}")

                    if "current_over_ball" in new_data:
                        logger.debug(f"Calling ball timeline update with: {repr(new_data['current_over_ball'])}")
                        update_ball_timeline_from_cov(new_data["current_over_ball"])


                    matched = True
                    break
            if not matched:
                logger.warning(f"[/update] Warning: Unexpected data at position {cursor}: {ble_string[cursor:cursor+10]}")
                cursor += 1
    else:
        new_data = incoming

    logger.info(f"[/update] Parsed PCS data: {new_data}")

    # Add the new parsed data to the pending buffer
    with pending_lock:
        pending_pcs_updates.append(new_data)

    # Handle flush timer
    if buffer_flush_timer is not None:
        buffer_flush_timer.cancel()

    buffer_flush_timer = threading.Timer(QUIET_TIME_SECONDS, process_pending_updates)
    buffer_flush_timer.start()

    return jsonify({"status": "success", "pending_updates": len(pending_pcs_updates)})






@app.route('/manual', methods=['POST'])
def manual_update():
    new_data = request.json

    init_data_file(DATA_MANUAL_FILE, read_json(DATA_PCS_FILE))
    manual_data = read_json(DATA_MANUAL_FILE)

    # Check for changes
    changed = False
    for key, value in new_data.items():
        if manual_data.get(key) != value:
            changed = True
            manual_data[key] = value

    if changed:
        manual_data["last_updated_at"] = datetime.utcnow().isoformat() + "Z"
        write_json(DATA_MANUAL_FILE, manual_data)
        logger.info(f"/manual endpoint - Data changed, writing to file: {new_data}")
        push_score_to_arduino()

# debugging line, but it fills the log
#    else:
#        logger.debug("/manual endpoint - Data unchanged, skipping write")

    # Always return current manual data (even if unchanged)
    return jsonify({"status": "success", "data": manual_data})

@app.route('/priority', methods=['GET'])
def get_priority():
    return jsonify(read_json(PRIORITY_FILE))
    
def get_current_priority():
    try:
        data = read_json(PRIORITY_FILE)
        return data.get("active_source", "Manual")  # fallback to Manual if key missing
    except Exception as e:
        logger.warning(f"Failed to read priority file: {e}")
        return "Manual"


@app.route('/set_priority', methods=['POST'])
def set_priority():
    data = request.json
    new_source = data.get("active_source")
    if new_source not in ["PCS", "Manual"]:
        return jsonify({"status": "error", "message": "Invalid source"}), 400
    priority = {"active_source": new_source, "updated_at": datetime.utcnow().isoformat() + "Z"}
    write_json(PRIORITY_FILE, priority)
    logger.info(f"/set_priority endpoint - Mode set to: {priority}")
    push_score_to_arduino()
    return jsonify({"status": "success", "active_source": new_source})

@app.route('/pcs_status')
def pcs_status():
    priority = read_json(PRIORITY_FILE)
    mode = priority.get("active_source", "PCS")

    pcs_data = read_json(DATA_PCS_FILE)
    manual_data = read_json(DATA_MANUAL_FILE)

    pcs_total = pcs_data.get("batting_team_score", "0/0")
    pcs_overs = pcs_data.get("overs_bowled", "0")
    pcs_timestamp = pcs_data.get("last_updated_at", "")
    startup_reset = pcs_data.get("startup_reset", False)

    manual_timestamp = manual_data.get("last_updated_at", "")
    pcs_total_val = pcs_total.split("/")[0] if "/" in pcs_total else pcs_total
    pcs_wickets = pcs_total.split("/")[1] if "/" in pcs_total else "0"

    show_pcs_warning = False
    pcs_ago = ""


    # Check if PCS data is newer than manual data and show a warning if appropriate.
    # This runs only when the current mode is "Manual" and the app is not in a startup reset state,
    #  startup reset state being the first boot of the day where data is reset.
    # It compares the PCS and manual timestamps:
    # - If PCS is more recent, set `show_pcs_warning = True` to alert the user that newer PCS data is available.
    # - If manual is newer (or equal), no warning is shown.
    # - If manual timestamp is missing, assume PCS is newer and show the warning.
    # Also logs the comparison result or any timestamp parsing errors for debugging.

    if pcs_timestamp and not startup_reset and mode == "Manual":
        try:
            pcs_dt = datetime.fromisoformat(pcs_timestamp.rstrip("Z"))
            pcs_ago = f"{(datetime.utcnow() - pcs_dt).seconds // 60}:{(datetime.utcnow() - pcs_dt).seconds % 60:02}"

            if manual_timestamp:
                manual_dt = datetime.fromisoformat(manual_timestamp.rstrip("Z"))
                if pcs_dt > manual_dt:
                    show_pcs_warning = True
                    logger.info(f"[PCS Warning] PCS data is newer than manual. PCS: {pcs_dt}, Manual: {manual_dt}")
                else:
                    logger.info(f"[PCS Check] PCS data is older or same as manual. PCS: {pcs_dt}, Manual: {manual_dt}")
            else:
                show_pcs_warning = True
                logger.info(f"[PCS Warning] Manual data has no timestamp, showing PCS warning. PCS: {pcs_dt}")

        except Exception as e:
            logger.warning(f"[PCS Warning] Timestamp parse error: {e}")
            pcs_ago = "?"

#    logger.debug(f"[PCS Status] Warning: {show_pcs_warning}, Mode: {mode}, PCS ago: {pcs_ago}")

    return jsonify({
        "show_pcs_warning": show_pcs_warning,
        "pcs_total": pcs_total_val,
        "pcs_wickets": pcs_wickets,
        "pcs_overs": pcs_overs,
        "pcs_last": pcs_timestamp,
        "pcs_ago": pcs_ago
    })


@app.route('/')
def index():
    return render_template('index.html')

@app.route('/manual')
def manual_update_page():
    priority = read_json(PRIORITY_FILE)
    mode = priority.get("active_source", "PCS")
    return render_template("manual.html", mode=mode, admin_password=repr(ADMIN_PASSWORD))


@app.route('/log_stream')
def log_stream():
    def generate():
        with open(LOG_FILE, 'r', encoding='utf-8', errors='replace') as f:
            # Read the last 200 lines first
            last_lines = deque(f, maxlen=200)
            for line in last_lines:
                yield f"data: {line.strip()}\n\n"
                time.sleep(0.01)  # tiny delay to prevent flood

            # Then tail live
            while True:
                line = f.readline()
                if line:
                    yield f"data: {line.strip()}\n\n"
                else:
                    time.sleep(0.2)

    return Response(stream_with_context(generate()), mimetype='text/event-stream')

@app.route("/serial_status")
def serial_status():
    if serial_mgr is None:
        return jsonify({
            "status": "Serial manager not initialized"
        })

    # Read current mode from priority file
    try:
        priority = read_json(PRIORITY_FILE)
        scoring_mode = priority.get("active_source", "Manual")
    except Exception as e:
        logger.warning(f"[serial_status] Failed to read priority file: {e}")
        scoring_mode = "Manual"

    # Extract values from serial_mgr
    status = {
        "ack_enabled": serial_mgr.ack_enabled,
        "connected": serial_mgr.ser is not None and serial_mgr.ser.is_open,
        "last_sent": serial_mgr.last_sent_msg or "",
        "last_ack": serial_mgr.last_ack_msg or "",
        "waiting_for_ack": serial_mgr.waiting_for_ack,
        "ack_timeout_exceeded": serial_mgr.ack_timeout_exceeded,
        "retry_count": serial_mgr.retry_count,
        "last_sent_values": serial_mgr.last_sent_values or [],
        "mode": scoring_mode,
        "ack_matched": None  # Only set if ACK matching is enabled
    }

    if serial_mgr.ack_enabled:
        status["ack_matched"] = serial_mgr.ack_matched_last

    return jsonify(status)

@app.route("/send_serial_command", methods=["POST"])
def send_serial_command():
    if serial_mgr is None:
        return jsonify({"status": "error", "message": "Serial manager not initialized"}), 500

    try:
        data = request.get_json()
        command = data.get("command", "").strip()
        if not command:
            return jsonify({"status": "error", "message": "No command provided"}), 400

        logger.info(f"[Admin] Sending raw command: {command}")
        response = serial_mgr.send_raw_command(command)
        return jsonify({"status": "success", "command": command, "response": response})

    except Exception as e:
        logger.error(f"[Admin] Failed to send raw serial command: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/admin')
def options_page():
    return render_template('admin.html')

@app.route("/check_admin_password", methods=["POST"])
def check_admin_password():
    data = request.get_json()
    provided = data.get("password", "")
    if provided == ADMIN_PASSWORD:
        return jsonify({"status": "ok"})
    return jsonify({"status": "fail"}), 403

    
@app.route('/save_options', methods=['POST'])
def save_options():
    data = request.json
    write_json(OPTIONS_FILE, data)

    # Also update matching fields in data_manual.json
    try:
        manual_data = read_json(DATA_MANUAL_FILE)
        if not isinstance(manual_data, dict):
            logger.info(f"/save_options endpoint - data empty")
            manual_data = {}

        for key in ['home_team', 'away_team', 'overs_per_innings']:
            if key in data:
                manual_data[key] = data[key]

        manual_data["last_updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        write_json(DATA_MANUAL_FILE, manual_data)
        logger.info(f"[/save_options] - Options saved: {data}")
        logger.info(f"[/save_options] - Options saved: {manual_data}")

    except Exception as e:
        logger.warning(f"/save_options endpoint - Could not sync Options to manual data: {e}")


    try:
        if firebase_mgr:
            firebase_mgr.publish("options", data)
            firebase_mgr.publish("data", manual_data)
    except Exception as e:
        logger.warning(f"[Firestore] Failed to push options: {e}")


    return jsonify({"status": "ok"})


@app.route('/load_options')
def load_options():
    if os.path.exists(OPTIONS_FILE):
        with open(OPTIONS_FILE, 'r') as f:
            return jsonify(json.load(f))
    return jsonify({})

@app.route('/upload_fixtures', methods=['POST'])
def upload_fixtures():
    file = request.files.get('file')
    if not file or not file.filename.endswith('.xlsx'):
        return jsonify({"status": "error", "message": "Invalid or missing file"}), 400

    wb = load_workbook(file, data_only=True)
    sheet = wb.active
    headers = [cell.value for cell in sheet[1]]
    fixtures = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))

        if row_data.get("Ground Owner") == HOME_GROUND_IDENTIFIER:
            excel_date = row_data.get("Date")
            if isinstance(excel_date, (int, float)):
                date_str = from_excel(excel_date).strftime("%d/%m/%Y")
            elif isinstance(excel_date, datetime):
                date_str = excel_date.strftime("%d/%m/%Y")
            else:
                date_str = str(excel_date)

            fixture = {
                "date": date_str,
                "home_team": row_data.get("Home Team", ""),
                "away_team": row_data.get("Away Team", ""),
                "type": row_data.get("Fixture Type", ""),
                "division_or_cup": row_data.get("Division / Cup", "")
            }

            fixtures.append(fixture)

    with open(FIXTURE_FILE, 'w') as f:
        json.dump(fixtures, f)
        logger.info(f"Fixtures uploaded: {fixtures}")

    return jsonify({"status": "success", "fixtures": fixtures})

@app.route('/fixtures')
def get_fixtures():
    try:
        if os.path.exists(FIXTURE_FILE):
            with open(FIXTURE_FILE, 'r') as f:
                fixtures = json.load(f)
            return jsonify(fixtures)
        else:
            return jsonify([])  # Return empty list if no file yet
    except Exception as e:
        logger.error(f"[Admin] Failed to load fixtures.json: {e}")
        return jsonify([]), 500

@app.route('/clear_fixtures', methods=['POST'])
def clear_fixtures():
    try:
        with open(FIXTURE_FILE, 'w') as f:
            json.dump([], f)
        logger.info("[Admin] Fixtures cleared by user.")
        return jsonify({"status": "success"})
    except Exception as e:
        logger.error(f"[Admin] Failed to clear fixtures: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/get_config')
def get_config():
    path = "/var/www/flaskapp/config.py"
    try:
        with open(path, "r", encoding="utf-8") as f:
            return f.read(), 200, {'Content-Type': 'text/plain; charset=utf-8'}
    except Exception as e:
        logger.error(f"[Admin] Failed to read config.py: {e}")
        return "Error loading config", 500


@app.route('/update_config', methods=['POST'])
def update_config():
    data = request.get_json()
    content = data.get('content', '')

    # Syntax check
    try:
        ast.parse(content)
    except SyntaxError as e:
        logger.warning(f"[Admin] Config syntax error: {e}")
        return jsonify({"status": "error", "message": f"Syntax error in config: {e}"}), 400

    # Runtime validation
    required_keys = ["SERIAL_OUTPUT_FORMAT", "CONFIG_TEMPLATE_VERSION"]
    config_scope = {}
    try:
        exec(content, config_scope)
        missing = [key for key in required_keys if key not in config_scope]
        if missing:
            logger.warning(f"[Admin] Missing keys in config: {missing}")
            return jsonify({"status": "error", "message": f"Missing required config fields: {missing}"}), 400
    except Exception as e:
        logger.warning(f"[Admin] Error validating config: {e}")
        return jsonify({"status": "error", "message": f"Runtime error: {e}"}), 400

    # Backup old config
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = f"/var/www/flaskapp/config_backup_{timestamp}.py"
        os.rename("/var/www/flaskapp/config.py", backup_path)
        logger.info(f"[Admin] config.py backed up as {backup_path}")
    except Exception as e:
        logger.warning(f"[Admin] Failed to backup config.py: {e}")

    # Write new config
    try:
        with open("/var/www/flaskapp/config.py", "w", encoding="utf-8") as f:
            f.write(content)
        logger.info("[Admin] config.py updated.")
    except Exception as e:
        logger.error(f"[Admin] Failed to write config.py: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

    # Restart Apache
    # Option A (safe way)
    try:
        subprocess.run(["touch", "/var/www/flaskapp/flaskapp.wsgi"], check=True)
        logger.info("[Admin] Triggered app reload via WSGI touch.")
    except Exception as e:
        logger.error(f"[Admin] Failed to trigger reload: {e}")
        return jsonify({"status": "error", "message": "Config saved, but failed to reload app."}), 500

    return jsonify({"status": "success"})



@app.route('/tv')
def tv_display():
    return render_template('tv.html')



@app.route('/shutdown', methods=['POST'])
def shutdown_pi():
    logger.warning("[Manual] Shutdown command triggered from web interface")
    try:
        subprocess.Popen(['sudo', 'shutdown', 'now'])
        return "OK"
    except Exception as e:
        logger.error(f"Failed to execute shutdown: {e}")
        return "FAIL", 500

@app.route('/reboot', methods=['POST'])
def reboot_pi():
    logger.warning("[Manual] Reboot command triggered from web interface")
    try:
        subprocess.Popen(['sudo', 'reboot'])
        return "OK"
    except Exception as e:
        logger.error(f"Failed to execute reboot: {e}")
        return "FAIL", 500

@app.route('/status')
def status():
    return jsonify({"status": "ok"})


@app.route('/client_heartbeat', methods=['POST'])
def client_heartbeat():
    ip = request.remote_addr
    now = time.time()
    data = request.get_json(force=True)


    # Update per-IP session info
    if ip not in client_sessions:
        client_sessions[ip] = {
            "first_seen": now,
            "last_seen": now,
            "pings": 1,
            "user_agent": data.get("userAgent", ""),
            "screen": data.get("screen", {}),
            "mode": data.get("mode", "unknown"),
            "overs": data.get("overs", 0),
            "innings": data.get("innings", 1),
            "ping_type_counts": {
                "initial": 0,
                "visible": 0,
                "interval": 0
            }
        }

    session = client_sessions[ip]
    session["last_seen"] = now
    session["pings"] += 1
    session["mode"] = data.get("mode", session["mode"])
    session["overs"] = data.get("overs", 0)
    session["innings"] = data.get("innings", 1)
    session["user_agent"] = data.get("userAgent", "")
    session["screen"] = data.get("screen", {})

    pt = data.get("ping_type", "interval")
    if pt in session["ping_type_counts"]:
        session["ping_type_counts"][pt] += 1

    # Append to per-ping log file (one line per ping)
    today_str = datetime.now().strftime("%Y-%m-%d")
    log_line = {
        "timestamp": now,
        "ip": ip,
        "ping_type": pt,
        "mode": session["mode"],
        "overs": session["overs"],
        "innings": session["innings"],
        "user_agent": session["user_agent"],
        "screen": session["screen"]
    }
    log_path = VIEWER_PINGS_LOG_TEMPLATE.format(date=today_str)
    try:
        with open(log_path, "a") as f:
            f.write(json.dumps(log_line) + "\n")
    except Exception as e:
        logger.warning(f"Failed to log viewer ping: {e}")

    return "OK"


@app.route('/viewer_data')
def viewer_data():
    today_str = datetime.now().strftime("%Y-%m-%d")
    path = VIEWER_SESSIONS_FILE_TEMPLATE.format(date=today_str)

    if os.path.exists(path):
        try:
            with open(path, "r") as f:
                sessions = json.load(f)
        except Exception as e:
            logger.warning(f"Failed to read viewer session file: {e}")
            sessions = {}
    else:
        sessions = {}

    return jsonify({ "sessions": sessions })

@app.route('/viewer_report')
def viewer_report_page():
    return render_template("viewer_report.html")

@app.route('/viewer_ping_data')
def viewer_ping_data():
    date_str = request.args.get("date")
    if not date_str:
        date_str = datetime.now().strftime("%Y-%m-%d")
    path = VIEWER_PINGS_LOG_TEMPLATE.format(date=date_str)
    result = []

    try:
        with open(path, "r") as f:
            for line in f:
                try:
                    entry = json.loads(line.strip())
                    result.append(entry)
                except:
                    continue
    except Exception as e:
        logger.warning(f"Failed to read viewer ping log: {e}")

    return jsonify(result)

@app.route('/viewer_ping_dates')
def viewer_ping_dates():
    folder = os.path.dirname(VIEWER_PINGS_LOG_TEMPLATE)
    prefix = os.path.basename(VIEWER_PINGS_LOG_TEMPLATE).split("{")[0]
    dates = []

    try:
        for f in os.listdir(folder):
            if f.startswith(prefix) and f.endswith(".jsonl"):
                date_part = f[len(prefix):-6]  # Strip prefix and '.jsonl'
                dates.append(date_part)
        dates.sort(reverse=True)
    except Exception as e:
        logger.warning(f"Failed to list ping log files: {e}")

    return jsonify(dates)


# ESP heartbeat


ESP_STATUS_FILE = '/var/www/flaskapp/esp_status.json'  # Adjust to your server paths

@app.route('/esp_heartbeat', methods=['POST'])
def esp_heartbeat():
    """Receive ESP32 heartbeat and save status."""
    data = request.get_json(force=True)
    logger.info(f"[ESP Heartbeat] - Heartbeat received: {data}")

    try:
        status = {
            "last_heartbeat": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "ble_connected": data.get("ble_connected", False),
        }
        write_json(ESP_STATUS_FILE, status)
    except Exception as e:
        logger.warning(f"[ESP Heartbeat] - Failed to save ESP status: {e}")

    return "OK", 200

@app.route('/esp_heartbeat', methods=['GET'])
def get_esp_status():
    """Provide latest ESP32 heartbeat status."""
    try:
        data = read_json(ESP_STATUS_FILE)
#        logger.info(f"[ESP Heartbeat] - Heartbeat data sent: {data}")
        return jsonify(data)
    except Exception as e:
        logger.warning(f"[ESP Heartbeat] - Failed to read ESP status: {e}")
        return jsonify({"error": "unavailable"}), 500


@app.route("/reset_data", methods=["POST"])
def reset_data():
    """Reset data files to default values from Admin page, same as startup routine."""

    logger.info("[Admin] Data files reset triggered from Admin page.")
    try:
        # Reset data_manual.json
        reset_manual = default_data.copy()
        reset_manual["last_updated_at"] = datetime.utcnow().isoformat() + "Z"
        write_json(DATA_MANUAL_FILE, reset_manual)

        # Reset data_pcs.json
        reset_pcs = default_data.copy()
        reset_pcs["last_updated_at"] = datetime.utcnow().isoformat() + "Z"
        reset_pcs["startup_reset"] = True  # Mark startup reset
        write_json(DATA_PCS_FILE, reset_pcs)

        # Set default priority to PCS
        write_json(PRIORITY_FILE, {
            "active_source": "PCS",
            "updated_at": datetime.utcnow().isoformat() + "Z"
        })

        # Apply today's fixture info
        auto_select_today_fixture()

        # Pause briefly to allow serial to stabilize
        time.sleep(2.0)

        # Push default scoreboard values
        push_score_to_arduino()

        logger.info("[Admin] Data files reset successfully via Admin page.")

        return jsonify({"status": "success", "message": "Data files reset."}), 200
    except Exception as e:
        logger.error(f"[Admin] Failed to reset data files: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500



##########################################################################
#   keep the startup routine at the end
##########################################################################


def startup_routine():

    logger.info(f"[Startup] Starting server.py with PID {os.getpid()}")


    global serial_mgr, firebase_mgr  # ensure you can set the module-level variable

    # Always start SerialManager
    serial_mgr = SerialManager(port=SERIAL_PORT, baudrate=SERIAL_BAUDRATE)
    logger.info("[Startup] SerialManager initialized")
    atexit.register(serial_mgr.stop)


    if FIREBASE_ENABLED:
        try:
            init_firebase(FIREBASE_CREDENTIAL_PATH)  # ✅ ensure Firebase is ready before anything else
            firebase_config = {
                "FIREBASE_ENABLED": FIREBASE_ENABLED,
                "FIREBASE_CREDENTIAL_PATH": FIREBASE_CREDENTIAL_PATH,
                "FIREBASE_PROJECT_ID": FIREBASE_PROJECT_ID,
                "FIREBASE_COLLECTION": FIREBASE_COLLECTION
            }
            firebase_mgr = FirebasePublisher(firebase_config)
            logger.info("[Startup] FirebaseManager initialized")
            atexit.register(firebase_mgr.stop)

  
        except Exception as e:
            logger.error(f"[Startup] Firebase initialization failed: {e}")
            firebase_mgr = None  # Optional fallback
  
        if firebase_mgr:
            from firebase_manager import delayed_push_viewer_date
            Thread(target=delayed_push_viewer_date, daemon=True).start()


    startup_flag = "/tmp/scoreboard_startup_done.flag"

    # Skip the rest if already initialized today
    if os.path.exists(startup_flag):
        logger.info("[Startup] Startup routine already completed - skipping.")
        return


    logger.info("[Startup] Checking scoreboard state on boot")

    # Check for freshness in PCS and Manual data
    is_pcs_fresh = is_data_fresh(DATA_PCS_FILE)
    is_manual_fresh = is_data_fresh(DATA_MANUAL_FILE)
    logger.info(f"[Startup] PCS data from today?: {is_pcs_fresh}, Manual data from today?: {is_manual_fresh}")

    if not is_pcs_fresh and not is_manual_fresh:
        logger.info("[Startup] No data found from today - treating as new day, defaulting data files, set mode to PCS")

        # Reset data using default_data
        reset_manual = default_data.copy()
        reset_manual["last_updated_at"] = datetime.utcnow().isoformat() + "Z"
        write_json(DATA_MANUAL_FILE, reset_manual)
      
        reset_pcs = default_data.copy()
        reset_pcs["last_updated_at"] = datetime.utcnow().isoformat() + "Z"
        reset_pcs["startup_reset"] = True  #  Only in PCS file
        write_json(DATA_PCS_FILE, reset_pcs)

        # Set mode to PCS - default so that no manual intervention needed to start scoring with PCS
        # this is only done first boot of the day, reboot with fresh existing data will retain mode already set
        write_json(PRIORITY_FILE, {
            "active_source": "PCS",
            "updated_at": datetime.utcnow().isoformat() + "Z"
        })


        # Apply fixture info (teams, overs logic)
        auto_select_today_fixture()

        time.sleep(2.0)  # to be sure serial has initialised
        # Push initial values to scoreboard
        push_score_to_arduino()
    else:
        logger.info("[Startup] Found data from today - no reset required")



    # Push current values to scoreboard
    push_score_to_arduino()

    try:
        with open(startup_flag, "w") as f:
            f.write(datetime.utcnow().isoformat())
        logger.info("[Startup] Startup routine complete - flag written to /run to prevent other threads repeating startup.")
    except Exception as e:
        logger.warning(f"[Startup] Failed to write startup flag: {e}")


############################################
# startup call to execute at Apache startup

try:
    startup_routine()
except Exception as e:
    logger.exception(f"[Startup] Failed to run startup_routine: {e}")




####################################################################
#   only use below if mot running in wsgi

#if __name__ == '__main__':
#    app.run(debug=True)
